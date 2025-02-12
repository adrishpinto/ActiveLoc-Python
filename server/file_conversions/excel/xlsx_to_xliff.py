import openpyxl
from lxml import etree

def xlsx_to_xliff(input_xlsx, output_xliff):
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active 
    
    root = etree.Element("xliff", version="1.2", xmlns="urn:oasis:names:tc:xliff:document:1.2")
    file_elem = etree.SubElement(root, "file", original=input_xlsx, datatype="spreadsheet", sourceLanguage="en")
    body = etree.SubElement(file_elem, "body")
    
    row_id = 1 
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        for col_id, cell in enumerate(row, start=1):
            if cell:  
                trans_unit = etree.SubElement(body, "trans-unit", id=f"{row_id}-{col_id}")
                source = etree.SubElement(trans_unit, "source")
                source.text = str(cell)  
    
    tree = etree.ElementTree(root)
    tree.write(output_xliff, pretty_print=True, xml_declaration=True, encoding="utf-8")
    print(f"XLIFF saved to {output_xliff}")


xlsx_to_xliff("sample.xlsx", "output.xliff")